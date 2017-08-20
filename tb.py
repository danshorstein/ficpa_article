import openpyxl

wb = openpyxl.load_workbook(r'excel\pbc_tb.xlsx', data_only=True)
sheet = wb.get_sheet_by_name('tb')
last_row = sheet.max_row
import_tb = [['gl', 'description', 'net']]

for row in range(4, last_row + 1):
    row = str(row)   
    gl_description = sheet['B'+row].value
    if gl_description == 'Total':
        continue
    gl, *description = gl_description.split(' - ')
    description = ' '.join(description)
    net = round(sheet['G'+row].value - sheet['H'+row].value, 0)
    import_tb.append([gl, description, net])

rounding = sum(row[2] for row in import_tb[1:])

if rounding != 1:
    import_tb.append(['9999', 'rounding', -rounding])

import_sheet = wb.create_sheet(title=r'tb_import', index=0)

for num, (gl, desc, net) in enumerate(import_tb):
    import_sheet['A' + str(num + 1)].value = gl
    import_sheet['B' + str(num + 1)].value = desc
    import_sheet['C' + str(num + 1)].value = net

wb.save(r'excel\tb_import.xlsx')
